"""
CACYOF OYSCATECH ALUMNI
Executive Records Collation and Categorization System
"""

import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

INPUT_FOLDER = "input_files"
OUTPUT_FOLDER = "Output"

# Position normalization map: raw fragment → canonical name
# Keys are lowercased substrings to match against raw position text
NORMALIZATION = {
    # ── President ────────────────────────────────────────────
    "president": "President",
    # ── Vice President ───────────────────────────────────────
    "vice president": "Vice President",
    # ── General Secretary ────────────────────────────────────
    "general secretary": "General Secretary",
    "gen. sec": "General Secretary",
    "gen sec": "General Secretary",
    # ── Assistant General Secretary ──────────────────────────
    "assistant general secretary": "Assistant General Secretary",
    "asst. gen. sec": "Assistant General Secretary",
    "asst gen sec": "Assistant General Secretary",
    "asst gen. sec": "Assistant General Secretary",
    # ── Prayer Coordinator ───────────────────────────────────
    "prayer coordinator": "Prayer Coordinator",
    "prayer cord": "Prayer Coordinator",
    "prayer co-ord": "Prayer Coordinator",
    "prayer": "Prayer Coordinator",          # catch bare "Prayer"
    # ── Assistant Prayer Coordinator ────────────────────────
    "assistant prayer coordinator": "Assistant Prayer Coordinator",
    "ass. prayer coordinator": "Assistant Prayer Coordinator",
    "asst. prayer coordinator": "Assistant Prayer Coordinator",
    "ass prayer": "Assistant Prayer Coordinator",
    "asst prayer": "Assistant Prayer Coordinator",
    "ass. prayer": "Assistant Prayer Coordinator",
    # ── Bible Coordinator ────────────────────────────────────
    "bible coordinator": "Bible Coordinator",
    "bible cord": "Bible Coordinator",
    "bible (cord)": "Bible Coordinator",
    "bible co-ord": "Bible Coordinator",
    "bible": "Bible Coordinator",            # catch bare "Bible"
    # ── Assistant Bible Coordinator ─────────────────────────
    "assistant bible coordinator": "Assistant Bible Coordinator",
    "ass. bible coordinator": "Assistant Bible Coordinator",
    "asst. bible coordinator": "Assistant Bible Coordinator",
    "assistant bible": "Assistant Bible Coordinator",
    "ass. bible": "Assistant Bible Coordinator",
    "asst. bible": "Assistant Bible Coordinator",
    # ── Brother Coordinator ──────────────────────────────────
    "brother coordinator": "Brother Coordinator",
    "brother cord": "Brother Coordinator",
    "brother co-ord": "Brother Coordinator",
    "brother": "Brother Coordinator",
    # ── Sister Coordinator ───────────────────────────────────
    "sister coordinator": "Sister Coordinator",
    "sister cord": "Sister Coordinator",
    "sisters coordinator": "Sister Coordinator",
    "sisters' coordinator": "Sister Coordinator",
    "sister co-ord": "Sister Coordinator",
    "sister 2": "Sister Coordinator",
    # ── Assistant Sister Coordinator ────────────────────────
    "assistant sister coordinator": "Assistant Sister Coordinator",
    "assistant sisters coordinator": "Assistant Sister Coordinator",
    "ass. sister coordinator": "Assistant Sister Coordinator",
    "asst. sister coordinator": "Assistant Sister Coordinator",
    "assistant sister": "Assistant Sister Coordinator",
    # ── Welfare Coordinator ──────────────────────────────────
    "welfare coordinator": "Welfare Coordinator",
    "welfare cord": "Welfare Coordinator",
    "welfare 1": "Welfare Coordinator",
    "welfare 2": "Welfare Coordinator",
    "welfare co-ord": "Welfare Coordinator",
    "welfare": "Welfare Coordinator",
    # ── Assistant Welfare Coordinator ───────────────────────
    "assistant welfare coordinator": "Assistant Welfare Coordinator",
    "ass. welfare coordinator": "Assistant Welfare Coordinator",
    "asst. welfare": "Assistant Welfare Coordinator",
    "asst welfare": "Assistant Welfare Coordinator",
    "ass welfare": "Assistant Welfare Coordinator",
    # ── Choir Coordinator ────────────────────────────────────
    "choir coordinator": "Choir Coordinator",
    "choir cord": "Choir Coordinator",
    "choir co-ord": "Choir Coordinator",
    "choir coordinator 2": "Choir Coordinator",
    "choir 2": "Choir Coordinator",
    # ── Assistant Choir Coordinator ─────────────────────────
    "assistant choir coordinator": "Assistant Choir Coordinator",
    "ass. choir coordinator": "Assistant Choir Coordinator",
    "asst choir": "Assistant Choir Coordinator",
    "assistant choir cord": "Assistant Choir Coordinator",
    "asst choir cord": "Assistant Choir Coordinator",
    # ── Drama Coordinator ────────────────────────────────────
    "drama coordinator": "Drama Coordinator",
    "drama cord": "Drama Coordinator",
    "drama co-ord": "Drama Coordinator",
    "drama coordinator 1": "Drama Coordinator",
    # ── Assistant Drama Coordinator ─────────────────────────
    "assistant drama coordinator": "Assistant Drama Coordinator",
    "ass. drama coordinator": "Assistant Drama Coordinator",
    "asst. drama coordinator": "Assistant Drama Coordinator",
    "assistant drama": "Assistant Drama Coordinator",
    "asst drama": "Assistant Drama Coordinator",
    # ── Technical Coordinator ────────────────────────────────
    "technical coordinator": "Technical Coordinator",
    "tech cord": "Technical Coordinator",
    "tech co-ord": "Technical Coordinator",
    "drama/tech cord": "Technical Coordinator",  # split via / first, this catches leftover
    # ── Assistant Technical Coordinator ─────────────────────
    "asst. technical": "Assistant Technical Coordinator",
    "asst technical": "Assistant Technical Coordinator",
    "assistant technical": "Assistant Technical Coordinator",
    "assistant drama/tech 2": "Assistant Drama / Technical Coordinator",
    # ── Usher Coordinator ────────────────────────────────────
    "usher coordinator": "Usher Coordinator",
    "usher cord": "Usher Coordinator",
    "usher co-ord": "Usher Coordinator",
    "chief usher": "Usher Coordinator",
    "usher coordinator 2": "Usher Coordinator",
    "usher 2": "Usher Coordinator",
    # ── Assistant Usher Coordinator ─────────────────────────
    "assistant usher coordinator": "Assistant Usher Coordinator",
    "asst. usher coordinator": "Assistant Usher Coordinator",
    "asst usher": "Assistant Usher Coordinator",
    "asst. usher": "Assistant Usher Coordinator",
    # ── Evangelism Coordinator ───────────────────────────────
    "evangelism coordinator": "Evangelism Coordinator",
    "evangelism cord": "Evangelism Coordinator",
    "evangelism": "Evangelism Coordinator",
    # ── Assistant Evangelism Coordinator ────────────────────
    "assistant evangelism coordinator": "Assistant Evangelism Coordinator",
    "ass. evangelism coordinator": "Assistant Evangelism Coordinator",
    "asst. evangelism coordinator": "Assistant Evangelism Coordinator",
    "assistant evangelism": "Assistant Evangelism Coordinator",
    "asst evangelism": "Assistant Evangelism Coordinator",
    # ── Academic Coordinator ─────────────────────────────────
    "academic coordinator": "Academic Coordinator",
    "academic cord": "Academic Coordinator",
    "academic coordinator 2": "Academic Coordinator",
    "sisters' academic coordinator": "Academic Coordinator",
    "sisters academic coordinator": "Academic Coordinator",
    "academic co-ord": "Academic Coordinator",
    "academic": "Academic Coordinator",
    # ── Assistant Academic Coordinator ──────────────────────
    "assistant academic coordinator": "Assistant Academic Coordinator",
    "asst. academic": "Assistant Academic Coordinator",
    "asst academic": "Assistant Academic Coordinator",
    "assistant academic": "Assistant Academic Coordinator",
    # ── Visitation Coordinator ───────────────────────────────
    "visitation coordinator": "Visitation Coordinator",
    "visitation cord": "Visitation Coordinator",
    "visitation (cord)": "Visitation Coordinator",
    "visitation": "Visitation Coordinator",
    # ── Assistant Visitation Coordinator ────────────────────
    "assistant visitation coordinator": "Assistant Visitation Coordinator",
    "asst. visitation": "Assistant Visitation Coordinator",
    "asst visitation": "Assistant Visitation Coordinator",
    "assistant visitation": "Assistant Visitation Coordinator",
    # ── Sanctuary Coordinator ────────────────────────────────
    "sanctuary coordinator": "Sanctuary Coordinator",
    "sanctuary cord": "Sanctuary Coordinator",
    # ── Financial Secretary ──────────────────────────────────
    "financial secretary": "Financial Secretary",
    # ── Treasurer ────────────────────────────────────────────
    "treasurer": "Treasurer",
    # ── PRO / Public Relations ───────────────────────────────
    "p.r.o": "PRO (Public Relations Officer)",
    " pro": "PRO (Public Relations Officer)",
    "/pro": "PRO (Public Relations Officer)",
    "/ pro": "PRO (Public Relations Officer)",
    "pro ": "PRO (Public Relations Officer)",
    # ── Decoration & Sanitation ──────────────────────────────
    "decoration": "Decoration / Sanitation Coordinator",
    "sanitation": "Decoration / Sanitation Coordinator",
    # ── Media Coordinator ────────────────────────────────────
    "media coordinator": "Media Coordinator",
    "media": "Media Coordinator",
    # ── Building Coordinator ─────────────────────────────────
    "building coordinator": "Building Coordinator",
    "building": "Building Coordinator",
}

# Longer/more specific keys should match before shorter ones.
# We sort by length descending to handle that.
SORTED_NORM_KEYS = sorted(NORMALIZATION.keys(), key=len, reverse=True)


def normalize_position(raw: str) -> list[str]:
    """
    Split a compound position string into individual canonical position names.
    Handles slash-separated roles, ampersand combos, and assistant prefixes.
    """
    if not isinstance(raw, str) or not raw.strip():
        return []

    # Pre-normalize known shorthand combos before splitting
    pre_map = {
        r"drama/tech\s*cord": "Drama Coordinator / Technical Coordinator",
        r"drama\s*/\s*tech\s*cord": "Drama Coordinator / Technical Coordinator",
        r"assistant\s+drama/tech\s*2?": "Assistant Drama Coordinator / Assistant Technical Coordinator",
        r"asst\.\s*choir\s*/\s*asst\.\s*academic": "Assistant Choir Coordinator / Assistant Academic Coordinator",
        r"asst\.\s*usher\s*/\s*asst\.\s*visitation": "Assistant Usher Coordinator / Assistant Visitation Coordinator",
        r"decoration\s*/\s*usher\s*2?": "Decoration / Sanitation Coordinator / Usher Coordinator",
        r"welfare/treasurer": "Welfare Coordinator / Treasurer",
        r"vice\s*president/prayer": "Vice President / Prayer Coordinator",
        r"academic/visitation\s*\(?cord\)?": "Academic Coordinator / Visitation Coordinator",
        r"drama/tech": "Drama Coordinator / Technical Coordinator",
        r"building\s*/\s*brother": "Building Coordinator / Brother Coordinator",
        r"prayer\s*/\s*drama": "Prayer Coordinator / Drama Coordinator",
        r"bible\s*/\s*academic": "Bible Coordinator / Academic Coordinator",
        r"usher\s*/\s*visitation": "Usher Coordinator / Visitation Coordinator",
        r"welfare\s*/\s*treasurer": "Welfare Coordinator / Treasurer",
        r"asst\.\s*welfare\s*/\s*financial\s*secretary": "Assistant Welfare Coordinator / Financial Secretary",
        r"financial\s*secretary\s*/\s*welfare": "Financial Secretary / Welfare Coordinator",
        r"media\s*/\s*assistant\s*prayer": "Media Coordinator / Assistant Prayer Coordinator",
        r"evangelism\s*/\s*visitation": "Evangelism Coordinator / Visitation Coordinator",
        r"assistant\s*evangelism\s*/\s*assistant\s*visitation": "Assistant Evangelism Coordinator / Assistant Visitation Coordinator",
        r"welfare\s*1\s*/\s*assistant\s*sisters?\s*coordinator": "Welfare Coordinator / Assistant Sister Coordinator",
        r"p\.r\.o\s*/\s*assistant\s*general\s*secretary": "PRO (Public Relations Officer) / Assistant General Secretary",
        r"asst\.\s*gen\.\s*sec\.\s*/\s*pro": "Assistant General Secretary / PRO (Public Relations Officer)",
        r"assistant\s*gen\s*sec\s*/\s*p\.r\.o": "Assistant General Secretary / PRO (Public Relations Officer)",
        r"ass\.\s*general\s*secretary\s*/\s*pro": "Assistant General Secretary / PRO (Public Relations Officer)",
        r"technical\s*coordinator\s*/\s*p\.r\.o": "Technical Coordinator / PRO (Public Relations Officer)",
        r"brother\s*/\s*bible\s*coordinator": "Brother Coordinator / Bible Coordinator",
        r"brother\s*/\s*prayer\s*coordinator": "Brother Coordinator / Prayer Coordinator",
        r"vice\s*president\s*/\s*bible\s*coordinator": "Vice President / Bible Coordinator",
        r"vice\s*president\s*/\s*prayer\s*coordinator": "Vice President / Prayer Coordinator",
        r"vice\s*president\s*/\s*treasurer": "Vice President / Treasurer",
        r"assistant\s*general\s*secretary\s*/\s*p\.r\.o": "Assistant General Secretary / PRO (Public Relations Officer)",
        r"sister\s*coordinator\s*/\s*choir\s*coordinator": "Sister Coordinator / Choir Coordinator",
        r"assistant\s*sister\s*coordinator\s*/\s*choir\s*coordinator": "Assistant Sister Coordinator / Choir Coordinator",
        r"assistant\s*prayer\s*coordinator\s*/\s*visitation\s*coordinator": "Assistant Prayer Coordinator / Visitation Coordinator",
        r"usher\s*coordinator\s*/\s*visitation\s*coordinator": "Usher Coordinator / Visitation Coordinator",
        r"welfare\s*coordinator\s*/\s*chief\s*usher": "Welfare Coordinator / Usher Coordinator",
    }

    processed = raw.strip()
    for pattern, replacement in pre_map.items():
        processed = re.sub(pattern, replacement, processed, flags=re.IGNORECASE)

    # Split on delimiters: ' / ', ' & ', ' and '
    parts = re.split(r"\s*/\s*|\s*&\s*|\s+and\s+", processed, flags=re.IGNORECASE)
    results = []

    for part in parts:
        part = part.strip().rstrip(".,;")
        if not part:
            continue
        lower = part.lower()

        # Special: bare "PRO" / "P.R.O" after split
        if lower.strip() in ("pro", "p.r.o"):
            canonical = "PRO (Public Relations Officer)"
            if canonical not in results:
                results.append(canonical)
            continue

        matched = None
        for key in SORTED_NORM_KEYS:
            if key in lower:
                matched = NORMALIZATION[key]
                break
        if matched:
            if matched not in results:
                results.append(matched)
        else:
            canonical = part.title()
            if canonical not in results:
                results.append(canonical)

    return results


def extract_session_year(raw_year_str: str) -> str:
    """Convert '2022/2023 SESSION' → '2022/2023'"""
    if not isinstance(raw_year_str, str):
        return str(raw_year_str)
    return raw_year_str.replace("SESSION", "").strip()


def read_file(filepath: str) -> list[dict]:
    """
    Read one Excel file and return a list of dicts:
    {name, positions: [str], phone, year}
    """
    df = pd.read_excel(filepath, header=None)

    # Row 0 is the year/session string
    year_raw = str(df.iloc[0, 1]) if pd.notna(df.iloc[0, 1]) else str(df.iloc[0, 0])
    year = extract_session_year(year_raw)

    # Detect columns by scanning header row (row 1)
    header_row = [str(v).lower() if pd.notna(v) else "" for v in df.iloc[1]]
    name_col = next((i for i, v in enumerate(header_row) if "name" in v), None)
    pos_col = next((i for i, v in enumerate(header_row) if "position" in v), None)
    phone_col = next((i for i, v in enumerate(header_row) if "phone" in v or "number" in v), None)

    if name_col is None or pos_col is None:
        print(f"  WARNING: Could not detect columns in {filepath}")
        return []

    records = []
    for idx in range(2, len(df)):
        row = df.iloc[idx]
        name = str(row.iloc[name_col]).strip() if pd.notna(row.iloc[name_col]) else ""
        if not name or name in ("nan", "NaN", "None"):
            continue

        # Collect all position values from pos_col onwards (some files have 2 position columns)
        raw_positions = []
        for ci in range(pos_col, min(pos_col + 3, len(row))):
            val = row.iloc[ci]
            if pd.notna(val) and str(val).strip() not in ("nan", "NaN", "None", ""):
                raw_positions.append(str(val).strip())

        combined = " / ".join(raw_positions)
        positions = normalize_position(combined)

        phone = ""
        if phone_col is not None:
            pval = row.iloc[phone_col]
            if pd.notna(pval) and str(pval).strip() not in ("nan", "NaN", "None", ""):
                phone = str(pval).strip()

        if positions:
            records.append({
                "name": name,
                "positions": positions,
                "phone": phone,
                "year": year,
            })

    return records


def build_position_map(all_records: list[dict]) -> dict[str, list[dict]]:
    """Group records by each individual position."""
    pos_map: dict[str, list[dict]] = {}
    for rec in all_records:
        for pos in rec["positions"]:
            pos_map.setdefault(pos, []).append({
                "name": rec["name"],
                "phone": rec["phone"],
                "year": rec["year"],
            })
    return pos_map


def safe_filename(pos: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", pos).replace(" ", "_")


# ─────────────────────────────────────────────
# STYLING
# ─────────────────────────────────────────────

HEADER_BG = "1F4E79"      # Deep navy
HEADER_FG = "FFFFFF"      # White
TITLE_BG  = "2E75B6"      # Mid blue
SUB_BG    = "D6E4F0"      # Light blue
ROW_ALT   = "EBF5FB"      # Very light blue alt row

thin = Side(border_style="thin", color="BFBFBF")
med  = Side(border_style="medium", color="1F4E79")

def cell_border(top=thin, bottom=thin, left=thin, right=thin):
    return Border(top=top, bottom=bottom, left=left, right=right)


def write_position_file(pos_name: str, rows: list[dict], out_dir: str):
    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r'[\\/*?:"<>\[\]|]', "-", pos_name)[:31]

    # ── Title rows ──────────────────────────────
    ws.merge_cells("A1:D1")
    t1 = ws["A1"]
    t1.value = "CACYOF OYSCATECH ALUMNI"
    t1.font = Font(name="Arial", size=14, bold=True, color=HEADER_FG)
    t1.fill = PatternFill("solid", fgColor=HEADER_BG)
    t1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:D2")
    t2 = ws["A2"]
    t2.value = pos_name.upper()
    t2.font = Font(name="Arial", size=12, bold=True, color=HEADER_FG)
    t2.fill = PatternFill("solid", fgColor=TITLE_BG)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── Column headers ───────────────────────────
    headers = ["S/N", "Name", "Year Served", "Phone Number"]
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=hdr)
        cell.font = Font(name="Arial", size=10, bold=True, color=HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border(top=med, bottom=med)
    ws.row_dimensions[3].height = 18

    # ── Data rows ────────────────────────────────
    for i, rec in enumerate(rows, 1):
        row_num = i + 3
        bg = PatternFill("solid", fgColor=ROW_ALT) if i % 2 == 0 else None

        for col, val in enumerate([i, rec["name"], rec["year"], rec["phone"]], 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left" if col == 2 else "center",
                                       vertical="center", wrap_text=True)
            cell.border = cell_border()
            if bg:
                cell.fill = bg
        ws.row_dimensions[row_num].height = 16

    # ── Column widths ────────────────────────────
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16

    # ── Save ─────────────────────────────────────
    fname = safe_filename(pos_name) + ".xlsx"
    fpath = os.path.join(out_dir, fname)
    wb.save(fpath)
    print(f"  ✔  {fname}  ({len(rows)} records)")


# ─────────────────────────────────────────────
# SUMMARY workbook (all positions on one sheet)
# ─────────────────────────────────────────────

def write_summary(pos_map: dict[str, list[dict]], out_dir: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "All Executives"

    # Title
    ws.merge_cells("A1:E1")
    t1 = ws["A1"]
    t1.value = "CACYOF OYSCATECH ALUMNI – COMPLETE EXECUTIVE RECORDS"
    t1.font = Font(name="Arial", size=13, bold=True, color=HEADER_FG)
    t1.fill = PatternFill("solid", fgColor=HEADER_BG)
    t1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    headers = ["S/N", "Name", "Position", "Year Served", "Phone Number"]
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = Font(name="Arial", size=10, bold=True, color=HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=TITLE_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border(top=med, bottom=med)
    ws.row_dimensions[2].height = 18

    global_sn = 1
    for pos_name in sorted(pos_map.keys()):
        for rec in pos_map[pos_name]:
            row_num = global_sn + 2
            bg = PatternFill("solid", fgColor=ROW_ALT) if global_sn % 2 == 0 else None
            vals = [global_sn, rec["name"], pos_name, rec["year"], rec["phone"]]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(
                    horizontal="left" if col in (2, 3) else "center",
                    vertical="center", wrap_text=True)
                cell.border = cell_border()
                if bg:
                    cell.fill = bg
            ws.row_dimensions[row_num].height = 16
            global_sn += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    fpath = os.path.join(out_dir, "_SUMMARY_ALL_EXECUTIVES.xlsx")
    wb.save(fpath)
    print(f"  ✔  _SUMMARY_ALL_EXECUTIVES.xlsx  ({global_sn - 1} total records)")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main(input_folder: str = INPUT_FOLDER, output_folder: str = OUTPUT_FOLDER):
    os.makedirs(output_folder, exist_ok=True)

    files = sorted([
        os.path.join(input_folder, f)
        for f in os.listdir(input_folder)
        if f.endswith(".xlsx") and not f.startswith("~")
    ])

    if not files:
        print(f"No .xlsx files found in '{input_folder}'")
        return

    print(f"\nReading {len(files)} file(s)...")
    all_records = []
    for fp in files:
        print(f"  → {os.path.basename(fp)}")
        recs = read_file(fp)
        all_records.extend(recs)
        print(f"     {len(recs)} executives extracted")

    pos_map = build_position_map(all_records)
    print(f"\nFound {len(pos_map)} unique positions across all files.\n")

    print(f"Generating output files in '{output_folder}/'...\n")
    for pos_name in sorted(pos_map.keys()):
        rows = pos_map[pos_name]
        write_position_file(pos_name, rows, output_folder)

    print()
    write_summary(pos_map, output_folder)

    print(f"\n✅  Done! {len(pos_map) + 1} files written to '{output_folder}/'")


if __name__ == "__main__":
    main()
